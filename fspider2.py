import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import time
from random import Random
import random
import time
def getJsid():
    url = 'http://www.soucase.com'

    se = requests.Session()
    se.get(url)
    js = se.cookies.get('JSESSIONID')
    return js

class Internet:

    def __init__(self):
        self.username = '02512578466'
        self.passwd = '683772'
    def outLine(self):
        cmd_str0 = "rasdial/DISCONNECT"
        not_found = os.system(cmd_str0)

    def onLine(self):
        cmd_str1 = "rasdial" + " " + "宽带连接" + " " + self.username + " " + self.passwd
        not_found = os.system(cmd_str1)
def changeip():
    line = Internet()
    line.outLine()
    time.sleep(1)
    line.onLine()
    time.sleep(5)
def createUser():
    str = ''
    chars = 'AaBbCcDdEeFfGgHhIiJjKkLlMmNnOoPpQqRrSsTtUuVvWwXxYyZz'
    length = len(chars) - 1
    random = Random()
    for i in range(8):
        str += chars[random.randint(0, length)]
    return str
def random_str():
    a = random.randint(1000,3500)
    print(a)
    return str(a)
userid = '00000'
username = 'username'
js = getJsid()
def getPage(url):
    session = requests.Session()
    global js
    global username
    global userid

    head = {
            'Accept': 'image / gif,image / jpeg, image / pjpeg, application / x - ms - application, application / xaml + xml, application / x - ms - xbap, * / *',
            'Accept-Encoding': 'gzip,deflate',
            'Accept-Language': 'zh-CN',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            "Cookie": 'userAuth=0;username_cssp='+username+';userid_cssp='+userid+';S=655EB60E885623A4DE64ABEDBA2EF3F9;JSESSIONID='+js,
            'User-Agent': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729)'
        }
    html = session.get(url=url, headers=head, allow_redirects=False)
    beObj = BeautifulSoup(html.text, "lxml")
    print(username,'usernam')
    print(userid,'userid')
    print(js,'js')

    return beObj
def urlCtrl(url,page):
    s = {}
    s1 = url.split('&')
    for i in s1:
        s2 = i.split('=')
        s[s2[0]] = s2[1]
    url2 = '?sw=' + s['?sw'] + '&' + \
           'field=' + s['field'] + '&' + \
           'selectAdv=' + s['selectAdv'] + '&' + \
           'swarrayAdv=' + s['swarrayAdv'] + '&' + \
           'fieldAdv=' + s['fieldAdv'] + '&' + \
           'orderID=' + s['orderID'] + '&' + \
           'orderType=' + s['orderType'] + '&' + \
           'orderStatus=' + s['orderStatus'] + '&' + \
           'channelid=' + s['channelid'] + '&' + \
           'pages=' + page + '&' + \
           'C_CASETYPE_CODE=' + s['C_CASETYPE_CODE'] + '&' + \
           'C_COUNT_CODE=' + s['C_COUNT_CODE'] + '&' + \
           'C_ORIGIN_PASS_CODE=' + s['C_ORIGIN_PASS_CODE'] + '&' + \
           'syear=' + s['syear']
    return url2



def urlCtrodown(url,page):
    s = {}
    s1 = url.split('&')
    for i in s1:
        s2 = i.split('=')
        s[s2[0]] = s2[1]
    url2 = '?sw=' + s['?sw'] + '&' + \
           'field=' + s['field'] + '&' + \
           'selectAdv=' + s['selectAdv'] + '&' + \
           'swarrayAdv=' + s['swarrayAdv'] + '&' + \
           'fieldAdv=' + s['fieldAdv'] + '&' + \
           'orderType=' + s['orderType'] + '&' + \
           'orderStatus=' + s['orderStatus'] + '&' + \
           'channelid=' + s['channelid'] + '&' + \
           'pages=' + page + '&' + \
           'C_CASETYPE_CODE=' + s['C_CASETYPE_CODE'] + '&' + \
           'C_COUNT_CODE=' + s['C_COUNT_CODE'] + '&' + \
           'C_ORIGIN_PASS_CODE=' + s['C_ORIGIN_PASS_CODE'] + '&' + \
           'syear=' + s['syear']+'&orderID=512&orderType=string&orderStatus=true&orderUpDown=down'
    return url2
def urlCtroup(url,page):
    s = {}
    s1 = url.split('&')
    for i in s1:
        s2 = i.split('=')
        s[s2[0]] = s2[1]
    url2 = '?sw=' + s['?sw'] + '&' + \
           'field=' + s['field'] + '&' + \
           'selectAdv=' + s['selectAdv'] + '&' + \
           'swarrayAdv=' + s['swarrayAdv'] + '&' + \
           'fieldAdv=' + s['fieldAdv'] + '&' + \
           'orderType=' + s['orderType'] + '&' + \
           'orderStatus=' + s['orderStatus'] + '&' + \
           'channelid=' + s['channelid'] + '&' + \
           'pages=' + page + '&' + \
           'C_CASETYPE_CODE=' + s['C_CASETYPE_CODE'] + '&' + \
           'C_COUNT_CODE=' + s['C_COUNT_CODE'] + '&' + \
           'C_ORIGIN_PASS_CODE=' + s['C_ORIGIN_PASS_CODE'] + '&' + \
           'syear=' + s['syear']+'&orderID=512&orderType=string&orderStatus=false&orderUpDown=up'
    return url2

'http://search.soucase.com/search.do?sw=&selectAdv=&swarrayAdv=&fieldAdv=&channelid=281&field=0&=&C_CASETYPE_CODE=1&C_COUNT_CODE=21&C_ORIGIN_PASS_CODE=1&syear=2014&orderID=512&orderType=string&orderStatus=true&orderUpDown=down'
def searchForYear(year):
    url = 'http://search.soucase.com/search.do?sw=&selectAdv=&swarrayAdv=&fieldAdv=&orderID=-1&orderType=&orderStatus=&channelid=281&field=0&syear='+year+'&C_CASETYPE_CODE=1&C_ORIGIN_PASS_CODE=1'
    # 2016年
    beObj = getPage(url)
    province = beObj.find_all("a",{"class":"leftList"})
    province.remove(province[10])
    return province

def spidForPage(url):
    global userid
    global username
    global js
    num = 0
    sb = getData(url)
    while sb[0]==[]:
        username = createUser()
        userid = random_str()
        js = getJsid()
        time.sleep(5)
        sb = getData(url)
        num = num +1
        if num==20:
            print('======================break mather fucker=========================')
            pass
    if sb[0]==[]:
        pass
    else:
        dealWithdata(sb[0],sb[1],sb[2])
        deleteline('url14.txt')
def getData(url):
    # a = getPage(url).find_all('td',{'class':'name'})
    # for i in a:
    #     url3 = url2+i.find_all('a')[1]['href']
    #     loadUrl(url3)
    # print(txturl)

    txturl = url.replace('detail','getDetailAllText')
    print(txturl)

    beObj1 = getPage(txturl)
    txt = str(beObj1.find('body'))#.replace('','')
    try:
        txt = txt.replace('<p>','').replace('</p>','\n').replace('<body>{"alltextValue":"','').replace('</body>','').replace('"}','')
    except:
        txt = beObj1.get_text()

    sazl = re.search(r"(?i)zl.{14}",txt)
    if sazl!=None:
        sazl = sazl.group()
    else:
        sazl = ""

    beObj = getPage(url)
    data = beObj.find_all('dl',{'class':'tab_con'})
    sb = [data,sazl,txt]
    print('===========================')
    return sb
def dealWithdata(data,sazl,txt):
    with open('1.txt', 'r+') as f:
        row = int(f.readline())
        print(row)
    s = {}
    data = data[0].get_text()+data[1].get_text()+data[2].get_text()

    try:
        data = data.replace('案件信息','').replace('法院信息','').replace('当事人信息','')
    except:
        pass
    data = data.split('\n')
    for i in data:
        s2 = i.split('：')
        try:
            s[s2[0]] = s2[1]
        except:
            pass
    print(s)
    wb = load_workbook(filename='2014.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    ws.cell(row=row,column=1).value = row
    ws.cell(row=row,column=2).value = s.get('案件号')
    ws.cell(row=row,column=3).value = s.get('案件名称')
    ws.cell(row=row,column=4).value = s.get('判决书类型')
    ws.cell(row=row,column=5).value = s.get('案由')
    ws.cell(row=row,column=6).value = s.get('审判金额')
    ws.cell(row=row,column=7).value = s.get('审结日期')
    ws.cell(row=row,column=8).value = s.get('受理日期')
    ws.cell(row=row,column=9).value = s.get('法院')
    ws.cell(row=row,column=10).value = s.get('法院级别')
    ws.cell(row=row,column=11).value = s.get('原告（上诉人）')
    ws.cell(row=row,column=12).value = s.get('被告（被诉人）')
    ws.cell(row=row,column=13).value = s.get('原告（上诉人）代表人')
    ws.cell(row=row,column=14).value = s.get('被告（被诉人）代表人')
    ws.cell(row=row,column=15).value = s.get('原告（上诉人）代理人')
    ws.cell(row=row,column=16).value = s.get('被告（被诉人）代理人')
    ws.cell(row=row,column=17).value = s.get('原告（上诉人）代理机构')
    ws.cell(row=row,column=18).value = s.get('被告（被诉人）代理机构')
    ws.cell(row=row,column=19).value = sazl

    wb.save(filename='2014.xlsx')
    with open('1.txt', 'w+') as f:
        sa= str(row+1)
        f.write(sa)
        print(row)
    try:
        with open('2014/'+s['案件号']+'.txt','w+') as f:
            f.writelines(txt)
    except:
        pass
    time.sleep(5)
def loadUrl(url):
    with open('url14.txt','a') as f :
        f.write(url+'\n')
def deleteline(txt):
    with open(txt, 'r') as old_file:
        with open(txt, 'r+') as new_file:
            current_line = 0
            while current_line < (1 - 1):
                old_file.readline()
                current_line += 1
            seek_point = old_file.tell()
            new_file.seek(seek_point, 0)
            a = old_file.readline()
            next_line = old_file.readline()
            while next_line:
                new_file.write(next_line)
                next_line = old_file.readline()
            new_file.truncate()
def readLineold(txt):
    with open(txt, 'r') as old_file:
        a = old_file.readline()
        return a
if __name__ == '__main__':
    try:
        while True:
            url = readLineold('url14.txt')
            spidForPage(url)
    except:
        time.sleep(20)
        # prov = searchForYear('2014')
        # url2 = 'http://search.soucase.com/search.do'
        # for p in prov:
        #
        #     ptxt = p.get_text()
        #     print(ptxt)
        #     num = int(ptxt[ptxt.find('(')+1:ptxt.find(')')])
        #     if num <200:
        #             if num%10!=0:
        #                 totalPage = int(num/10+1)
        #             else:
        #                 totalPage = int(num/10)
        #             for page in range(totalPage):
        #                 page =str(page+1)
        #                 urllist = urlCtrl(p['href'], page)
        #                 url = url2 + urllist
        #                 print(url)
        #                 #loadUrl(url)
        #                 spidForPage(url)
        #     else:
        #             if num%10!=0:
        #                 totalPage = int(num/10+1)
        #             else:
        #                 totalPage = int(num/10)
        #             for page in range(totalPage):
        #                 page =page+1
        #                 if page < 20:
        #                     pag = str(page + 1)
        #                     urldown = url2+urlCtrodown(p['href'], str(page))
        #                     print(urldown)
        #                     #loadUrl(urldown)
        #                     spidForPage(urldown)
        #                 else:
        #                     pag = str(page + 1 - 20)
        #                     urlup = url2+urlCtroup(p['href'], str(page))
        #                     spidForPage(urlup)
        #                     print(urlup)
        #                     #loadUrl(urlup)

